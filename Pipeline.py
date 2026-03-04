"""
Pipeline.py
===========
Main entry point for the STM self-citation analysis pipeline.

Usage
-----
  python pipeline.py <input1.xlsx> [<input2.xlsx> ...] --output <output.xlsx>

Arguments
---------
  input files   One or more Weekly Status Excel workbooks (each must contain
                a "Request Status" sheet and a "Chapter Author" sheet).
  --output      Path for the output Excel file.
                Default: STM_Output_<timestamp>.xlsx next to the first input file.
  --verbose     Print progress messages.

Output format
-------------
  One pair of sheets per rights-holder with total assets > 30 (after dedup):
    "<Rights Holder> by chapter"  — Net Rights Holder Assets + yellow sum row
    "<Rights Holder> by source"   — Count sum row, orange highlight for Count > 3
  The "Rights Holder" column is omitted (redundant with sheet name).

Architecture
------------
.. code-block:: text

   Pipeline.py                     ← CLI entry-point & Excel writer
       ├── STM_Core.py             ← shared pre-processing & name matching
       ├── STM_By_Chapter.py       ← chapter-level aggregation
       └── STM_By_Source.py        ← source-level grouping

Dependencies
------------
  pandas  == 3.0.1
  numpy   == 2.4.2
  openpyxl >= 3.1
"""

import argparse
import sys
from datetime import datetime
from pathlib import Path

# Ensure sibling modules are importable when run via .venv\Scripts\python.exe
sys.path.insert(0, str(Path(__file__).resolve().parent))

import numpy as np             # noqa: E402  — used in vectorised column-width calc
import pandas as pd            # noqa: E402
from openpyxl.styles import PatternFill   # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from STM_By_Chapter import run_by_chapter  # noqa: E402
from STM_By_Source import run_by_source    # noqa: E402


# ─────────────────────────────────────────────────────────────────────────────
# Colour & threshold constants
# ─────────────────────────────────────────────────────────────────────────────

# openpyxl PatternFill objects are lightweight and reused across all cells.
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

# STM Permissions gratis limit per rights-holder per chapter / per source.
# Assets whose Net count exceeds this value are highlighted in orange.
STM_GRATIS_LIMIT = 3

# Minimum total assets a rights-holder must have across the book to be
# included in the output.  Rights-holders below this threshold are skipped
# entirely to keep the output file concise.
MIN_TOTAL_ASSETS = 30


# ─────────────────────────────────────────────────────────────────────────────
# Excel loading helpers
# ─────────────────────────────────────────────────────────────────────────────

PERM_SHEET = "Request Status"
AUTH_SHEET = "Chapter Author"


def load_workbook_data(path: Path, verbose: bool = False) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Load the two required sheets from a single input workbook.

    Parameters
    ----------
    path : Path
        Absolute path to an Excel (.xlsx) file that must contain exactly the
        sheets ``"Request Status"`` and ``"Chapter Author"``.
    verbose : bool
        If ``True``, print the filename being loaded.

    Returns
    -------
    tuple[pd.DataFrame, pd.DataFrame]
        ``(df_perm, df_auth)`` — the raw permissions DataFrame and the raw
        chapter-author DataFrame.

    Raises
    ------
    ValueError
        If one or both required sheets are missing from the workbook.  The
        error message names the missing sheet(s) explicitly so the user can
        fix the file without reading the traceback.

    Data Flow
    ---------
    +-----------+-------------------------------+------------------+
    | Input     | Process                       | Output           |
    +===========+===============================+==================+
    | path      | pd.read_excel(sheet_name=...) | df_perm (raw)    |
    +-----------+-------------------------------+------------------+
    | path      | pd.read_excel(sheet_name=...) | df_auth (raw)    |
    +-----------+-------------------------------+------------------+
    | df_perm   | dropna(axis=1, how="all")     | df_perm (pruned) |
    +-----------+-------------------------------+------------------+
    | df_auth   | dropna(axis=1, how="all")     | df_auth (pruned) |
    +-----------+-------------------------------+------------------+
    """
    if verbose:
        print(f"  Loading: {path.name}")

    # Validate sheet existence before reading to produce a clear error message.
    # pd.ExcelFile is cheaper than two failed pd.read_excel calls.
    try:
        xf = pd.ExcelFile(path)
    except Exception as exc:
        raise ValueError(f"Cannot open workbook '{path.name}': {exc}") from exc

    missing = [s for s in (PERM_SHEET, AUTH_SHEET) if s not in xf.sheet_names]
    if missing:
        raise ValueError(
            f"'{path.name}' is missing required sheet(s): {missing}. "
            f"Available sheets: {xf.sheet_names}"
        )

    df_perm = pd.read_excel(path, sheet_name=PERM_SHEET, header=0)
    df_auth = pd.read_excel(path, sheet_name=AUTH_SHEET, header=0)

    # Drop completely-empty columns that Excel sometimes adds as artefacts.
    df_perm = df_perm.dropna(axis=1, how="all")
    df_auth = df_auth.dropna(axis=1, how="all")

    return df_perm, df_auth


# ─────────────────────────────────────────────────────────────────────────────
# Per-sheet post-processing helpers
# ─────────────────────────────────────────────────────────────────────────────

# Column names used by the 'by chapter' sheet — must match run_by_chapter output.
_COL_TOTAL   = "Total Rights holder Assets"
_COL_SELF    = "Chapter Author is Source Author"
_COL_NET     = "Net Rights Holder Assets"
_COL_COUNT   = "Count"
_COL_RH      = "Rights Holder"


def _build_sum_row(df: pd.DataFrame, sum_cols: list[str]) -> pd.DataFrame:
    """
    Build a single-row sum-row DataFrame for appending to a results sheet.

    **Vectorised implementation**: replaces the original ``for col in df.columns``
    dict-building loop with direct pandas column selection and ``pd.Series``
    construction.

    Complexity
    ----------
    Old (loop): O(k) Python iterations over k columns.
    New (vectorised): O(1) Python overhead; NumPy handles the summation
    over ``sum_cols`` in a single C-level pass.

    Parameters
    ----------
    df : pd.DataFrame
        The data-rows-only DataFrame (must NOT already contain a sum row).
    sum_cols : list[str]
        Column names whose values should be summed.  All other columns get
        ``None`` (rendered as blank in Excel).

    Returns
    -------
    pd.DataFrame
        A one-row DataFrame with the same columns as ``df``.

    Data Flow
    ---------
    +---------------+--------------------------------------+-------------------+
    | Input         | Process                              | Output            |
    +===============+======================================+===================+
    | df, sum_cols  | df[sum_cols].sum() → pd.Series       | sums: Series      |
    +---------------+--------------------------------------+-------------------+
    | sums          | reindex(df.columns, fill_value=None) | sum_row: Series   |
    +---------------+--------------------------------------+-------------------+
    | sum_row       | pd.DataFrame([sum_row])              | 1-row pd.DataFrame|
    +---------------+--------------------------------------+-------------------+
    """
    # Filter sum_cols to only those that actually exist in df (defensive guard
    # against renamed/missing columns — avoids KeyError on edge-case inputs).
    valid_sum_cols = [c for c in sum_cols if c in df.columns]

    # Vectorised: compute sums for all requested columns in one pass.
    # reindex() fills all other columns with None so the row has the right shape.
    sums = df[valid_sum_cols].sum()
    sum_row_series = sums.reindex(df.columns, fill_value=None)

    return pd.DataFrame([sum_row_series])


def _prepare_chapter_sheet(df: pd.DataFrame) -> pd.DataFrame:
    """
    Prepare a 'by chapter' slice for Excel output.

    Transformations applied (in order):
      1. Drop ``Rights Holder`` column (redundant — it is the sheet name).
      2. Add ``Net Rights Holder Assets`` = ``Total Rights holder Assets``
         − ``Chapter Author is Source Author``.
      3. Append a vectorised sum row covering the two numeric columns.

    Parameters
    ----------
    df : pd.DataFrame
        One rights-holder's slice from ``run_by_chapter()`` output. Must
        contain ``Total Rights holder Assets`` and
        ``Chapter Author is Source Author`` columns.

    Returns
    -------
    pd.DataFrame
        The transformed DataFrame with the sum row appended as the last row.

    Edge Cases
    ----------
    * Missing ``_COL_TOTAL`` or ``_COL_SELF`` → net column will be all-NaN
      but no ``KeyError``; ``_build_sum_row`` guards against missing cols too.
    * Empty input ``df`` → returns a one-row sum-row DataFrame (all zeros).

    Data Flow
    ---------
    +--------------------+-----------------------------------+----------------------+
    | Input              | Process                           | Output               |
    +====================+===================================+======================+
    | df (raw slice)     | drop "Rights Holder"              | df (fewer cols)      |
    +--------------------+-----------------------------------+----------------------+
    | Total, Self cols   | vectorised subtraction            | "Net RH Assets" col  |
    +--------------------+-----------------------------------+----------------------+
    | df + sum_cols list | _build_sum_row() → 1-row frame    | sum_row: pd.DataFrame|
    +--------------------+-----------------------------------+----------------------+
    | df + sum_row       | pd.concat([df, sum_row])          | final: pd.DataFrame  |
    +--------------------+-----------------------------------+----------------------+
    """
    df = df.copy()

    # 1. Drop Rights Holder — the sheet tab name already conveys this.
    if _COL_RH in df.columns:
        df = df.drop(columns=[_COL_RH])

    # 2. Net Rights Holder Assets — vectorised column arithmetic (no loop).
    #    Guard: if either source column is absent, produce NaN gracefully.
    if _COL_TOTAL in df.columns and _COL_SELF in df.columns:
        df[_COL_NET] = df[_COL_TOTAL] - df[_COL_SELF]
    else:
        df[_COL_NET] = np.nan

    # 3. Append vectorised sum row.
    sum_row = _build_sum_row(df, sum_cols=[_COL_TOTAL, _COL_NET])
    return pd.concat([df, sum_row], ignore_index=True)


def _prepare_source_sheet(df: pd.DataFrame) -> pd.DataFrame:
    """
    Prepare a 'by source' slice for Excel output.

    Transformations applied (in order):
      1. Drop ``Rights Holder`` column.
      2. Append a vectorised sum row covering the ``Count`` column.

    Parameters
    ----------
    df : pd.DataFrame
        One rights-holder's slice from ``run_by_source()`` output.

    Returns
    -------
    pd.DataFrame
        The transformed DataFrame with the sum row appended as the last row.

    Edge Cases
    ----------
    * Missing ``Count`` column → ``_build_sum_row`` returns a zero-only row
      without raising a ``KeyError``.
    * Empty input → returns a one-row sum-of-zeros DataFrame.

    Data Flow
    ---------
    +--------------------+-----------------------------+----------------------+
    | Input              | Process                     | Output               |
    +====================+=============================+======================+
    | df (raw slice)     | drop "Rights Holder"        | df (fewer cols)      |
    +--------------------+-----------------------------+----------------------+
    | df + ["Count"]     | _build_sum_row()            | sum_row: pd.DataFrame|
    +--------------------+-----------------------------+----------------------+
    | df + sum_row       | pd.concat([df, sum_row])    | final: pd.DataFrame  |
    +--------------------+-----------------------------+----------------------+
    """
    df = df.copy()

    # 1. Drop Rights Holder.
    if _COL_RH in df.columns:
        df = df.drop(columns=[_COL_RH])

    # 2. Append vectorised sum row (Count column only).
    sum_row = _build_sum_row(df, sum_cols=[_COL_COUNT])
    return pd.concat([df, sum_row], ignore_index=True)


# ─────────────────────────────────────────────────────────────────────────────
# Excel writer helpers
# ─────────────────────────────────────────────────────────────────────────────

def _auto_col_width(ws, df: pd.DataFrame) -> None:
    """
    Auto-size worksheet columns to fit their widest cell value.

    Uses vectorised ``str.len()`` on each column (via pandas) to avoid a
    Python-level for-loop over every cell.  The column header length is also
    considered.  Width is capped at 82 characters to prevent excessively wide
    columns for long bibliographic source strings.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        The openpyxl worksheet to modify in-place.
    df : pd.DataFrame
        The DataFrame whose content was just written to ``ws``.

    Edge Cases
    ----------
    * All-NaN column → ``fillna("")`` produces an all-empty string Series;
      ``.str.len().max()`` returns 0, so falls back to the header length.
    * Empty DataFrame (0 rows) → ``max_len`` is just ``len(col_name)``.

    Data Flow
    ---------
    +-------------------+------------------------------------------+--------------------------------+
    | Input             | Process                                  | Output                         |
    +===================+==========================================+================================+
    | df[col].fillna("") | .astype(str).str.len()                 | int Series of per-cell widths  |
    +-------------------+------------------------------------------+--------------------------------+
    | int Series        | .max() → compare with len(header)        | max_len: int                   |
    +-------------------+------------------------------------------+--------------------------------+
    | max_len           | min(max_len + 2, 82) → column_dimensions | ws column width set            |
    +-------------------+------------------------------------------+--------------------------------+
    """
    for col_idx, col_name in enumerate(df.columns, start=1):
        # Vectorised: fillna → astype(str) → str.len() → max() runs in NumPy
        col_str = df[col_name].fillna("").astype(str)
        max_len = max(
            len(str(col_name)),
            col_str.str.len().max() if len(df) > 0 else 0,
        )
        # Guard against NaN from .max() on an all-NaN series after fillna → str
        if not isinstance(max_len, (int, float)) or np.isnan(max_len):
            max_len = len(str(col_name))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(int(max_len) + 2, 82)


def _apply_highlight_to_sum_row(ws, df: pd.DataFrame, fill: PatternFill) -> None:
    """
    Apply a fill colour to every cell in the last (sum) row of a worksheet
    that contains a non-null value.

    Factored out of the two formatting functions to eliminate a second
    code duplication within ``pipeline.py`` itself.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
    df : pd.DataFrame
        The DataFrame that was written to ``ws`` (including the sum row).
    fill : PatternFill
        The openpyxl fill to apply (e.g. ``YELLOW_FILL``).

    Data Flow
    ---------
    +----------------+-------------------------------------------+------------------------+
    | Input          | Process                                   | Output                 |
    +================+===========================================+========================+
    | df.iloc[-1]    | pd.notna() boolean mask                   | notna_mask: bool array |
    +----------------+-------------------------------------------+------------------------+
    | notna_mask     | enumerate() → ws.cell().fill = fill      | cells filled in ws     |
    +----------------+-------------------------------------------+------------------------+
    """
    # sum row is always the last DataFrame row; Excel row = len(df) + 1 (for header)
    sum_excel_row = len(df) + 1
    last_row_vals = df.iloc[-1]

    for col_idx, has_value in enumerate(pd.notna(last_row_vals), start=1):
        if has_value:
            ws.cell(row=sum_excel_row, column=col_idx).fill = fill


def _apply_orange_to_column(ws, df: pd.DataFrame, col_name: str, threshold: float) -> None:
    """
    Apply orange fill to all data cells (excluding the sum row) in a given
    column whose numeric value exceeds ``threshold``.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
    df : pd.DataFrame
        The DataFrame that was written to ``ws`` (including the sum row).
    col_name : str
        Name of the column to scan.  If not present in ``df``, this is a no-op.
    threshold : float
        Cells with values strictly greater than this number get orange fill.

    Edge Cases
    ----------
    * Column not present → silent no-op (no ``KeyError``).
    * Non-numeric cell value (e.g. a NaN or a stray string) → skipped via
      ``try/except`` to avoid crashing the entire write operation.

    Data Flow
    ---------
    +------------------------+------------------------------------------+----------------------------+
    | Input                  | Process                                  | Output                     |
    +========================+==========================================+============================+
    | df[col_name] index     | cols.index(col_name) + 1                 | net_col: int (Excel col)   |
    +------------------------+------------------------------------------+----------------------------+
    | data rows (exclude -1) | float(cell.value) > threshold            | cell.fill = ORANGE_FILL    |
    +------------------------+------------------------------------------+----------------------------+
    """
    cols = list(df.columns)
    if col_name not in cols:
        return

    col_excel_idx = cols.index(col_name) + 1  # openpyxl is 1-indexed
    # Data rows: Excel row 2 (after header) up to the row before the sum row.
    # len(df) includes the sum row; sum row is the last row, so data goes to
    # Excel row len(df) (= len(df) - 1 data rows + 1 header row).
    n_data_rows = len(df) - 1  # exclude the sum row

    for row_idx in range(2, 2 + n_data_rows):   # row 2 = first data row
        cell = ws.cell(row=row_idx, column=col_excel_idx)
        try:
            if cell.value is not None and float(cell.value) > threshold:
                cell.fill = ORANGE_FILL
        except (ValueError, TypeError):
            # Non-numeric cell — silently skip to avoid crashing the writer.
            pass


def _apply_chapter_formatting(ws, df: pd.DataFrame) -> None:
    """
    Apply conditional formatting to a 'by chapter' worksheet:

    - **Yellow** on all non-blank cells in the sum row (last row).
    - **Orange** on any ``Net Rights Holder Assets`` cell in the data rows
      whose value is greater than ``STM_GRATIS_LIMIT`` (= 3).

    This signals to the user that the STM "Rule of 3" has been breached.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
    df : pd.DataFrame
        The chapter DataFrame (including sum row) that was written to ``ws``.
    """
    # Yellow on the sum row (calls shared helper — no code duplication)
    _apply_highlight_to_sum_row(ws, df, YELLOW_FILL)
    # Orange on Net Rights Holder Assets > STM_GRATIS_LIMIT
    _apply_orange_to_column(ws, df, _COL_NET, STM_GRATIS_LIMIT)


def _apply_source_formatting(ws, df: pd.DataFrame) -> None:
    """
    Apply conditional formatting to a 'by source' worksheet:

    - **Yellow** on all non-blank cells in the sum row (last row).
    - **Orange** on any ``Count`` cell in the data rows whose value
      is greater than ``STM_GRATIS_LIMIT`` (= 3).

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
    df : pd.DataFrame
        The source DataFrame (including sum row) that was written to ``ws``.
    """
    _apply_highlight_to_sum_row(ws, df, YELLOW_FILL)
    _apply_orange_to_column(ws, df, _COL_COUNT, STM_GRATIS_LIMIT)


def write_output(
    results: dict[str, tuple[pd.DataFrame, pd.DataFrame]],
    output_path: Path,
    verbose: bool = False,
) -> None:
    """
    Write all result DataFrames into a single multi-sheet Excel workbook.

    For each rights-holder label in ``results``, two sheets are created:
    ``"<label> by chapter"`` and ``"<label> by source"``.  Both receive
    auto-sized column widths and the appropriate conditional formatting.

    Parameters
    ----------
    results : dict[str, tuple[pd.DataFrame, pd.DataFrame]]
        ``{ rh_label: (chapter_df_prepared, source_df_prepared) }``
        as returned by ``run_pipeline()``.
    output_path : Path
        Absolute path for the output ``.xlsx`` file.  Parent directories
        are created if they do not already exist.
    verbose : bool
        If ``True``, print one line per sheet written.

    Edge Cases
    ----------
    * ``results`` is empty → creates an empty workbook with no sheets (the
      caller should have already warned the user via the console balance output).
    """
    if verbose:
        print(f"\nWriting output → {output_path}")

    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for rh_label, (ch_df, src_df) in results.items():
            ch_sheet  = f"{rh_label} by chapter"
            src_sheet = f"{rh_label} by source"

            # ── Chapter sheet ─────────────────────────────────────────────────
            ch_df.to_excel(writer, sheet_name=ch_sheet, index=False)
            ws_ch = writer.sheets[ch_sheet]
            _auto_col_width(ws_ch, ch_df)
            _apply_chapter_formatting(ws_ch, ch_df)
            if verbose:
                print(f"  [{ch_sheet}]  {len(ch_df)-1} data rows")

            # ── Source sheet ──────────────────────────────────────────────────
            src_df.to_excel(writer, sheet_name=src_sheet, index=False)
            ws_src = writer.sheets[src_sheet]
            _auto_col_width(ws_src, src_df)
            _apply_source_formatting(ws_src, src_df)
            if verbose:
                print(f"  [{src_sheet}]  {len(src_df)-1} data rows")


# ─────────────────────────────────────────────────────────────────────────────
# Core pipeline logic
# ─────────────────────────────────────────────────────────────────────────────

def run_pipeline(
    input_paths: list[Path],
    output_path: Path,
    verbose: bool = False,
) -> dict[str, tuple[pd.DataFrame, pd.DataFrame]]:
    """
    Load all input files, run both analyses, filter by asset threshold,
    apply sum rows and Net column, run the balance check, and return a
    result dict ready for ``write_output()``.

    Parameters
    ----------
    input_paths : list[Path]
        One or more paths to Weekly Status ``.xlsx`` workbooks.
    output_path : Path
        Destination path for the output file (used for logging only here;
        actual writing is done by ``write_output()``).
    verbose : bool
        Print progress messages to stdout.

    Returns
    -------
    dict[str, tuple[pd.DataFrame, pd.DataFrame]]
        ``{ rh_label: (chapter_df_prepared, source_df_prepared) }``

    Raises
    ------
    FileNotFoundError
        If none of the provided input paths exist or yield valid data.
    SystemExit
        If all rights-holders are filtered out by ``MIN_TOTAL_ASSETS``.

    Data Flow (high-level)
    ---------
    +-------------------+-----------------------------+-------------------------+
    | Input             | Process                     | Output                  |
    +===================+=============================+=========================+
    | input_paths       | load_workbook_data()        | df_perm, df_auth (lists)|
    +-------------------+-----------------------------+-------------------------+
    | all_perm/auth     | pd.concat(ignore_index=True)| df_perm_all, df_auth_all|
    +-------------------+-----------------------------+-------------------------+
    | df_perm_all/auth  | run_by_chapter()            | chapter_df              |
    +-------------------+-----------------------------+-------------------------+
    | df_perm_all/auth  | run_by_source()             | source_df               |
    +-------------------+-----------------------------+-------------------------+
    | per-RH slices     | filter + prepare + balance  | results dict            |
    +-------------------+-----------------------------+-------------------------+
    """

    # ── 1. Load all input files ───────────────────────────────────────────────
    all_perm: list[pd.DataFrame] = []
    all_auth: list[pd.DataFrame] = []

    for path in input_paths:
        if not path.exists():
            print(f"[WARNING] File not found, skipping: {path}", file=sys.stderr)
            continue
        df_perm, df_auth = load_workbook_data(path, verbose=verbose)
        all_perm.append(df_perm)
        all_auth.append(df_auth)

    if not all_perm:
        raise FileNotFoundError("No valid input files were loaded.")

    # Concatenate across workbooks (supports multi-file inputs).
    df_perm_all = pd.concat(all_perm, ignore_index=True)
    df_auth_all = pd.concat(all_auth, ignore_index=True)

    if verbose:
        print(f"\nCombined rows → permissions: {len(df_perm_all)}, authors: {len(df_auth_all)}")

    # ── 2. Run both analyses ──────────────────────────────────────────────────
    if verbose:
        print("\nRunning 'by chapter' analysis...")
    chapter_df = run_by_chapter(df_perm_all, df_auth_all)

    if verbose:
        print("Running 'by source' analysis...")
    source_df = run_by_source(df_perm_all, df_auth_all)

    # ── 3. Collect all rights-holder names (union of both analyses) ───────────
    # Using a set-union so that a rights-holder that appears in one result
    # but not the other is still included in the processing loop.
    rights_holders = sorted(
        set(chapter_df["Rights Holder"].dropna().unique())
        | set(source_df["Rights Holder"].dropna().unique())
    )

    if verbose:
        print(f"\nRights-holders found ({len(rights_holders)}): {rights_holders}")

    # ── 4. Build per-rights-holder results ───────────────────────────────────
    results: dict[str, tuple[pd.DataFrame, pd.DataFrame]] = {}
    balance_ok = True

    print()
    print("  Rights-Holder Filter & Balance Check")
    print("  " + "-" * 56)

    for rh in rights_holders:
        # Slice each analysis result to this rights-holder only.
        ch_slice  = chapter_df[chapter_df["Rights Holder"] == rh].copy().reset_index(drop=True)
        src_slice = source_df[source_df["Rights Holder"] == rh].copy().reset_index(drop=True)

        # ── Enhancement 6: asset-count filter ────────────────────────────────
        # Skip rights-holders whose total asset count is ≤ MIN_TOTAL_ASSETS.
        # This avoids cluttering the output with small publishers.
        total_assets = int(ch_slice[_COL_TOTAL].sum()) if _COL_TOTAL in ch_slice.columns else 0
        if total_assets <= MIN_TOTAL_ASSETS:
            print(f"  ⏭  SKIPPED  {rh:<35s} (total assets = {total_assets} ≤ {MIN_TOTAL_ASSETS})")
            continue

        # ── Add Net column + sum rows (vectorised, see helpers above) ─────────
        ch_prepared  = _prepare_chapter_sheet(ch_slice)
        src_prepared = _prepare_source_sheet(src_slice)

        # ── Balance check ─────────────────────────────────────────────────────
        # The sum of Net Rights Holder Assets (chapter view) MUST equal the sum
        # of Count (source view) because both represent non-self-citation assets.
        # A mismatch indicates a bug in the deduplication or merge logic.
        # We exclude the last row (sum row) when computing these totals.
        net_sum   = int(ch_prepared[_COL_NET].iloc[:-1].sum())
        count_sum = int(src_prepared[_COL_COUNT].iloc[:-1].sum())
        balanced  = (net_sum == count_sum)

        if balanced:
            status = f"✅ BALANCED   net={net_sum}, count={count_sum}"
        else:
            status = f"⚠️  MISMATCH   net={net_sum} ≠ count={count_sum}"
            balance_ok = False

        print(f"  {status}  |  {rh}")

        # ── Sheet name (Excel tab label max 31 chars total) ───────────────────
        # The suffix " by chapter" / " by source" costs 11 chars, so labels
        # are capped at 20 chars.
        rh_label = rh[:20].strip() if len(rh) > 20 else rh
        results[rh_label] = (ch_prepared, src_prepared)

    print("  " + "-" * 56)
    if balance_ok:
        print("  ✅ All rights-holders pass the balance check.")
    else:
        print("  ⚠️  Some rights-holders have mismatches — review console output above.")
    print()

    if not results:
        print(
            "[WARNING] No rights-holders passed the asset-count filter "
            f"(MIN_TOTAL_ASSETS = {MIN_TOTAL_ASSETS}). "
            "Output file will be empty.",
            file=sys.stderr,
        )

    return results


# ─────────────────────────────────────────────────────────────────────────────
# CLI helpers
# ─────────────────────────────────────────────────────────────────────────────

def _default_output_path(input_paths: list[Path]) -> Path:
    """Generate a timestamped default output path next to the first input file."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return input_paths[0].parent / f"STM_{ts}.xlsx"


def main() -> None:
    """Parse CLI arguments and run the full pipeline."""
    parser = argparse.ArgumentParser(
        description="STM self-citation analysis pipeline",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "inputs", nargs="+", metavar="INPUT.xlsx",
        help="One or more Weekly Status Excel input files",
    )
    parser.add_argument(
        "--output", "-o", metavar="OUTPUT.xlsx",
        help="Path for the output Excel file (auto-generated if omitted)",
    )
    parser.add_argument(
        "--verbose", "-v", action="store_true",
        help="Print progress messages",
    )

    args = parser.parse_args()

    input_paths = [Path(p) for p in args.inputs]
    output_path = Path(args.output) if args.output else _default_output_path(input_paths)

    print("=" * 60)
    print("  STM Self-Citation Analysis Pipeline  (v3)")
    print("=" * 60)
    print(f"  Input files : {[p.name for p in input_paths]}")
    print(f"  Output      : {output_path}")
    print(f"  Asset filter: > {MIN_TOTAL_ASSETS} total assets")
    print(f"  STM limit   : > {STM_GRATIS_LIMIT} highlighted in orange")
    print()

    results = run_pipeline(input_paths, output_path, verbose=args.verbose)
    write_output(results, output_path, verbose=args.verbose)

    print()
    print("  ✅ Done!")
    print("=" * 60)


if __name__ == "__main__":
    main()
